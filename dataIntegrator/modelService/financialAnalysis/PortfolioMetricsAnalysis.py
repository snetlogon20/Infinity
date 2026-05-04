"""
投资组合指标分析工具（完全参考 stock_metrics_analysis.py）

计算并导出多个金融指标到 Excel 文件，包括：
- 均数 (Mean)
- 偏度 (Skewness)
- 峰度 (Kurtosis)
- 标准差 (Sigma)
- 与基准指数的相关系数
- Treynor Ratio (特雷诺比率)
- Sharpe Ratio (夏普比率)
- Information Ratio (信息比率)
- Sortino Ratio (索提诺比率/SOR)
- CML Weight (资本市场线最优权重)
"""

from dataIntegrator import CommonLib, CommonParameters
from dataIntegrator.common.CommonDataParameters import CommonDataParameters
from dataIntegrator.modelService.financialAnalysis.RiskFreeRateManager import RiskFreeRateManager
import pandas as pd
import numpy as np
import os
from datetime import datetime
from io import BytesIO
from statsmodels import regression
import statsmodels.api as sm
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import Series
from scipy.optimize import minimize

logger = CommonLib.logger


class PortfolioMetricsAnalysis:
    """投资组合指标分析器（完全参考 stock_metrics_analysis.py）"""

    def __init__(self, market_symbol='000001.SH', trading_days=252):
        self.market_symbol = market_symbol  # 市场基准指数
        self.trading_days = trading_days  # 年交易日数

    def get_stock_list(self, stock_type="cn_blue_chip"):
        """
        根据类型获取股票列表（参考 CMLAnalysisTest）
        
        参数:
        - stock_type: 股票类型
                      美股: ['us_tech', 'us_finance', 'us_mixed', 'us_custom']
                      A股: ['cn_blue_chip', 'cn_tech', 'cn_consumer', 'cn_financial', 'cn_energy', 'cn_custom']
        
        返回:
        - stocks: 股票代码列表
        - market_type: 市场类型 ('US' 或 'CN')
        - market_symbol: 市场指数符号
        """
        # 美国股票组合
        if stock_type == "us_tech":
            stocks = ["SPY", "AAPL", "MSFT", "NVDA", "GOOGL", "META", "TSLA", "AVGO", "ADBE"]
            market_type = "US"
            market_symbol = "SPY"

        elif stock_type == "us_finance":
            stocks = ["SPY", "C", "JPM", "GS", "MS", "BAC", "WFC", "BLK", "AXP"]
            market_type = "US"
            market_symbol = "SPY"

        elif stock_type == "us_mixed":
            stocks = [
                "SPY", "C", "JPM", "AAPL", "NVDA", "GS", "MS", "GE",
                "MSFT", "AVGO", "ADBE", "UNH", "JNJ", "LLY", "PFE", "MRK", "AMZN",
                "TSLA", "MCD", "NFLX", "HD", "GOOGL", "META", "DIS", "CMCSA", "T",
                "CAT", "UPS", "BA", "HON", "PG", "KO", "PEP", "WMT", "COST", "XOM",
                "CVX", "COP", "SLB", "EOG", "AMT", "PLD", "EQIX", "SPG", "O", "NEE",
                "DUK", "SO", "D", "EXC", "LIN", "APD", "FCX", "NEM", "SHW"
            ]
            market_type = "US"
            market_symbol = "SPY"

        elif stock_type == "us_custom":
            stocks = ["SPY", "C", "JPM", "AAPL", "NVDA", "MSFT"]
            market_type = "US"
            market_symbol = "SPY"

        # 中国A股组合
        elif stock_type == "cn_blue_chip":
            stocks = [
                '000001.SH',
                '600519.SH',
                '601318.SH',
                '600036.SH',
                '601012.SH',
                '000858.SZ',
                '000333.SZ',
                '600276.SH',
                '601888.SH',
            ]
            market_type = "CN"
            market_symbol = "000001.SH"

        elif stock_type == "cn_tech":
            stocks = [
                '399001.SZ',
                '002093.SZ',
                '000902.SZ',
                '688498.SH',
                '002475.SZ',
                '002594.SZ',
                '300750.SZ',
                '000063.SZ',
                '600745.SH',
            ]
            market_type = "CN"
            market_symbol = "399001.SZ"

        elif stock_type == "cn_consumer":
            stocks = [
                '000001.SH',
                '600519.SH',
                '000858.SZ',
                '000333.SZ',
                '600887.SH',
                '603288.SH',
                '002714.SZ',
                '600104.SH',
                '601888.SH',
            ]
            market_type = "CN"
            market_symbol = "000001.SH"

        elif stock_type == "cn_financial":
            stocks = [
                '000300.SH',
                '601318.SH',
                '600036.SH',
                '601398.SH',
                '601288.SH',
                '601166.SH',
                '600030.SH',
                '601688.SH',
                '000776.SZ',
            ]
            market_type = "CN"
            market_symbol = "000300.SH"

        elif stock_type == "cn_energy":
            stocks = [
                '000001.SH',
                '601012.SH',
                '600438.SH',
                '601899.SH',
                '600028.SH',
                '601857.SH',
                '601368.SH',
                '600490.SH',
                '600470.SH',
            ]
            market_type = "CN"
            market_symbol = "000001.SH"

        elif stock_type == "cn_custom":
            stocks = CommonDataParameters.STOCK_LIST.copy()
            stocks.insert(0, {'ts_code': '000001.SH', 'name': '上证指数'})
            stocks = [stock['ts_code'] if isinstance(stock, dict) else stock for stock in stocks]
            market_type = "CN"
            market_symbol = "000001.SH"

        else:
            raise ValueError(
                f"不支持的股票类型: {stock_type}。"
                f"支持的类型: ['us_tech', 'us_finance', 'us_mixed', 'us_custom', "
                f"'cn_blue_chip', 'cn_tech', 'cn_consumer', 'cn_financial', 'cn_energy', 'cn_custom']")

        return stocks, market_type, market_symbol

    def fetch_stock_data(self, stocks, start_date, end_date):
        """从 ClickHouse 获取股票数据"""
        logger.info(f"开始获取 {len(stocks)} 只股票的数据...")
        
        dfs = {}
        
        for stock in stocks:
            is_market_index = (stock == self.market_symbol)
            
            if is_market_index:
                sql = f"""
                SELECT
                    trade_date as trade_date,
                    close as close_point
                FROM df_tushare_cn_index_daily
                WHERE ts_code = '{stock}'
                  AND trade_date >= '{start_date}'
                  AND trade_date <= '{end_date}'
                ORDER BY trade_date ASC
                """
            else:
                sql = f"""
                SELECT
                    trade_date as trade_date,
                    close as close_point
                FROM df_tushare_stock_daily
                WHERE ts_code = '{stock}'
                  AND trade_date >= '{start_date}'
                  AND trade_date <= '{end_date}'
                ORDER BY trade_date ASC
                """
            
            from dataIntegrator.dataService.ClickhouseService import ClickhouseService
            clickhouseService = ClickhouseService()
            df = clickhouseService.getDataFrameWithoutColumnsName(sql)
            
            if df.empty:
                logger.warning(f"警告: {stock} 在 {start_date} 到 {end_date} 期间没有数据")
                continue
            
            df['trade_date'] = pd.to_datetime(df['trade_date'])
            df.set_index('trade_date', inplace=True)
            dfs[stock] = df
        
        logger.info(f"成功获取 {len(dfs)} 只股票的数据")
        return dfs

    def filter_common_dates(self, dfs):
        """过滤出所有股票共同的交易日期"""
        stock_keys = list(dfs.keys())
        if not stock_keys:
            raise ValueError("没有可用的股票数据")
        
        common_dates = set(dfs[stock_keys[0]].index)
        for key in stock_keys[1:]:
            if key in dfs:
                common_dates = common_dates.intersection(set(dfs[key].index))
        
        common_dates = sorted(list(common_dates))
        
        if len(common_dates) == 0:
            raise ValueError("错误：所有股票没有共同的交易日！")
        
        filtered_dfs = {}
        for key in stock_keys:
            if key in dfs:
                filtered_dfs[key] = dfs[key].loc[common_dates]
        
        logger.info(f"共同日期数量: {len(common_dates)}, 范围: {common_dates[0]} 至 {common_dates[-1]}")
        return filtered_dfs

    def calculate_daily_returns(self, dfs):
        """计算日收益率（对数收益率）"""
        logger.info("计算日收益率...")
        
        returns_df = pd.DataFrame()
        
        for stock, df in dfs.items():
            close_prices = df['close_point']
            daily_return = np.log(close_prices / close_prices.shift(1))
            returns_df[stock] = daily_return
        
        returns_df = returns_df.dropna()
        logger.info(f"收益率数据形状: {returns_df.shape}")
        
        return returns_df

    def calculate_beta(self, stock_returns, market_returns):
        """计算 β 系数"""
        aligned_data = pd.DataFrame({
            'stock': stock_returns,
            'market': market_returns
        }).dropna()
        
        if len(aligned_data) == 0:
            return np.nan
        
        x = aligned_data['market'].values
        y = aligned_data['stock'].values
        
        x = sm.add_constant(x)
        model = regression.linear_model.OLS(y, x).fit()
        return model.params[1]

    def calculate_tracking_error(self, stock_returns, market_returns):
        """计算跟踪误差（年化）"""
        active_returns = stock_returns - market_returns
        tracking_error = np.std(active_returns) * np.sqrt(self.trading_days)
        return tracking_error

    def calculate_cml_optimal_weights(self, stocks, returns_df, risk_free_rate):
        """
        根据 CML (资本市场线) 计算最优权重
        
        参数:
        - stocks: 股票列表
        - returns_df: 收益率 DataFrame
        - risk_free_rate: 无风险利率
        
        返回:
        - weights_dict: 字典 {stock: weight}
        """
        # 准备数据
        stock_returns = returns_df[stocks].dropna()
        n_assets = len(stocks)
        
        if n_assets == 0 or stock_returns.empty:
            return {stock: np.nan for stock in stocks}
        
        # 计算协方差矩阵和预期收益
        cov_matrix = stock_returns.cov() * self.trading_days
        mean_returns = stock_returns.mean() * self.trading_days
        
        # 目标函数：最大化夏普比率（等价于最小化负夏普比率）
        def neg_sharpe_ratio(weights):
            portfolio_return = np.sum(mean_returns * weights)
            portfolio_volatility = np.sqrt(np.dot(weights.T, np.dot(cov_matrix, weights)))
            
            if portfolio_volatility == 0:
                return 0
            
            sharpe = (portfolio_return - risk_free_rate) / portfolio_volatility
            return -sharpe  # 最小化负夏普比率
        
        # 约束条件：权重之和为 1
        constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
        
        # 边界条件：每个权重在 0 到 1 之间（不允许做空）
        bounds = tuple((0, 1) for _ in range(n_assets))
        
        # 初始权重：等权重
        initial_weights = np.array([1/n_assets] * n_assets)
        
        try:
            # 优化
            result = minimize(
                neg_sharpe_ratio,
                initial_weights,
                method='SLSQP',
                bounds=bounds,
                constraints=constraints,
                options={'maxiter': 1000, 'ftol': 1e-10}
            )
            
            if result.success:
                optimal_weights = result.x
                weights_dict = {stocks[i]: optimal_weights[i] for i in range(n_assets)}
                return weights_dict
            else:
                logger.warning(f"CML 优化失败: {result.message}")
                return {stock: np.nan for stock in stocks}
        
        except Exception as e:
            logger.warning(f"CML 权重计算异常: {str(e)}")
            return {stock: np.nan for stock in stocks}

    def calculate_downside_deviation(self, returns, target_return=0):
        """计算下行偏差（年化）"""
        downside_returns = returns[returns < target_return]
        
        if len(downside_returns) == 0:
            return 0.0
        
        downside_variance = np.mean((downside_returns - target_return) ** 2)
        downside_deviation = np.sqrt(downside_variance) * np.sqrt(self.trading_days)
        
        return downside_deviation

    def get_stock_name(self, ts_code):
        """获取股票名称"""
        if ts_code == self.market_symbol:
            return "上证指数"
        
        try:
            from dataIntegrator.dataService.ClickhouseService import ClickhouseService
            clickhouseService = ClickhouseService()
            sql = f"""
            SELECT name
            FROM df_tushare_stock_basic
            WHERE ts_code = '{ts_code}'
            LIMIT 1
            """
            df = clickhouseService.getDataFrameWithoutColumnsName(sql)
            
            if not df.empty:
                return df.iloc[0]['name']
            else:
                return ts_code
        except Exception as e:
            logger.warning(f"获取股票名称失败 {ts_code}: {e}")
            return ts_code

    def analyze_stocks_rolling(self, stock_type="cn_blue_chip", start_date_fixed=None, end_date_start=None, 
                                end_date_end=None, window_days=360, risk_free_rate=None):
        """
        执行滚动回测分析（完全参考 stock_metrics_analysis.py）
        
        参数:
        - stock_type: 股票类型
        - start_date_fixed: 固定的开始日期 (格式: 'YYYYMMDD')，窗口起始点
        - end_date_start: 滚动结束日期的起点 (格式: 'YYYYMMDD')
        - end_date_end: 滚动结束日期的终点 (格式: 'YYYYMMDD')，默认为今天
        - window_days: 回看窗口天数（默认360天）
        - risk_free_rate: 年化无风险利率（如果为None，则自动获取）
        
        返回:
        - all_results_df: 包含所有滚动窗口的结果 DataFrame
        """
        # 设置默认日期
        if start_date_fixed is None:
            start_date_fixed = CommonDataParameters.get_start_date(days=window_days + 360)
        if end_date_start is None:
            end_date_start = '20251201'
        if end_date_end is None:
            end_date_end = CommonParameters.today
        
        # 获取股票列表和市场信息
        stocks, market_type, market_symbol = self.get_stock_list(stock_type)
        
        # 更新市场指数
        self.market_symbol = market_symbol
        
        logger.info("=" * 80)
        logger.info("🚀 开始滚动回测分析")
        logger.info(f"   股票类型: {stock_type}")
        logger.info(f"   市场类型: {market_type}")
        logger.info(f"   市场指数: {market_symbol}")
        logger.info(f"   固定起始日期: {start_date_fixed}")
        logger.info(f"   滚动结束日期范围: {end_date_start} 至 {end_date_end}")
        logger.info(f"   回看窗口: {window_days} 天")
        logger.info(f"   股票数量: {len(stocks)}")
        logger.info("=" * 80)
        
        # 生成滚动日期序列（只取交易日）
        rolling_dates = self._generate_rolling_dates(end_date_start, end_date_end)
        logger.info(f"滚动窗口数量: {len(rolling_dates)}")
        
        all_results = []
        
        for idx, current_end_date in enumerate(rolling_dates, 1):
            # 计算当前窗口的起始日期
            current_start_date = self._calculate_window_start(current_end_date, window_days)
            
            logger.info(f"\n{'='*60}")
            logger.info(f" 滚动窗口 {idx}/{len(rolling_dates)}")
            logger.info(f"   窗口范围: {current_start_date} 至 {current_end_date}")
            logger.info(f"{'='*60}")
            
            try:
                # 执行单次分析
                results_df = self._analyze_single_window(
                    stocks, current_start_date, current_end_date, risk_free_rate
                )
                
                # 添加回测日期列（放在第一列）
                results_df.insert(0, '回测日期', current_end_date)
                
                all_results.append(results_df)
                
                logger.info(f"✅ 窗口 {idx} 完成，分析了 {len(results_df)} 只股票")
                
            except Exception as e:
                logger.error(f"❌ 窗口 {idx} ({current_end_date}) 分析失败: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                continue
        
        # 合并所有结果
        if all_results:
            all_results_df = pd.concat(all_results, ignore_index=True)
            
            # 排序：日期正序、Sharpe Ratio 倒序、CML Weight 倒序
            all_results_df = all_results_df.sort_values(
                by=['回测日期', 'Sharpe Ratio', 'CML Weight'],
                ascending=[True, False, False]
            ).reset_index(drop=True)
            
            logger.info(f"\n{'='*80}")
            logger.info("✅ 滚动回测完成！")
            logger.info(f"   总窗口数: {len(rolling_dates)}")
            logger.info(f"   总记录数: {len(all_results_df)}")
            logger.info(f"{'='*80}")
            return all_results_df
        else:
            raise ValueError("没有成功的滚动窗口分析结果")
    
    def _generate_rolling_dates(self, start_date, end_date):
        """
        生成滚动日期序列（只包含交易日）
        
        参数:
        - start_date: 起始日期 (格式: 'YYYYMMDD')
        - end_date: 结束日期 (格式: 'YYYYMMDD')
        
        返回:
        - dates: 交易日列表
        """
        from dataIntegrator.dataService.ClickhouseService import ClickhouseService
        
        # 获取市场指数的交易日历
        clickhouseService = ClickhouseService()
        sql = f"""
        SELECT DISTINCT trade_date
        FROM df_tushare_cn_index_daily
        WHERE ts_code = '{self.market_symbol}'
          AND trade_date >= '{start_date}'
          AND trade_date <= '{end_date}'
        ORDER BY trade_date ASC
        """
        df = clickhouseService.getDataFrameWithoutColumnsName(sql)
        
        if df.empty:
            raise ValueError(f"在 {start_date} 至 {end_date} 期间没有找到交易日")
        
        dates = df['trade_date'].tolist()
        logger.info(f"找到 {len(dates)} 个交易日用于滚动回测")
        
        return dates
    
    def _calculate_window_start(self, end_date, window_days):
        """
        根据结束日期和窗口天数计算起始日期
        
        参数:
        - end_date: 结束日期 (格式: 'YYYYMMDD')
        - window_days: 窗口天数
        
        返回:
        - start_date: 起始日期 (格式: 'YYYYMMDD')
        """
        from datetime import timedelta
        
        # 转换日期格式
        end_dt = datetime.strptime(end_date, '%Y%m%d')
        
        # 向前推算窗口天数（考虑周末，实际交易日会更少）
        start_dt = end_dt - timedelta(days=window_days)
        
        return start_dt.strftime('%Y%m%d')
    
    def _analyze_single_window(self, stocks, start_date, end_date, risk_free_rate=None):
        """
        执行单个窗口的分析（完全参考 stock_metrics_analysis.py）
        
        参数:
        - stocks: 股票列表
        - start_date: 开始日期
        - end_date: 结束日期
        - risk_free_rate: 无风险利率
        
        返回:
        - results_df: 分析结果 DataFrame
        """
        # 获取无风险利率
        if risk_free_rate is None:
            riskFreeRateManager = RiskFreeRateManager()
            risk_free_rate = riskFreeRateManager.get_risk_free_rate(
                start_date, end_date, interest_country='CN'
            )
        
        # Step 1: 获取数据
        dfs = self.fetch_stock_data(stocks, start_date, end_date)
        
        if not dfs:
            raise ValueError("未能获取任何股票数据")
        
        # Step 2: 过滤共同日期
        filtered_dfs = self.filter_common_dates(dfs)
        
        # Step 3: 计算收益率
        returns_df = self.calculate_daily_returns(filtered_dfs)
        
        # Step 4: 计算 CML 最优权重
        market_returns = returns_df[self.market_symbol]
        # 只使用在收益率数据框中存在的股票来计算CML权重
        available_stocks = [stock for stock in stocks if stock in returns_df.columns]
        cml_weights = self.calculate_cml_optimal_weights(available_stocks, returns_df, risk_free_rate)
        
        # Step 5: 计算各项指标
        results = []
        
        for stock in stocks:
            if stock not in returns_df.columns:
                continue
            
            stock_returns = returns_df[stock]
            
            # 基本统计量
            mean_annual = stock_returns.mean() * self.trading_days
            skewness = stock_returns.skew()
            kurtosis = stock_returns.kurtosis()
            sigma_annual = stock_returns.std() * np.sqrt(self.trading_days)
            
            # 与市场的相关系数
            correlation = stock_returns.corr(market_returns) if stock != self.market_symbol else 1.0
            
            # Beta 系数
            beta = self.calculate_beta(stock_returns, market_returns) if stock != self.market_symbol else 1.0
            
            # Treynor Ratio
            if beta != 0 and not np.isnan(beta):
                treynor_ratio = (mean_annual - risk_free_rate) / beta
            else:
                treynor_ratio = np.nan
            
            # Sharpe Ratio
            if sigma_annual != 0:
                sharpe_ratio = (mean_annual - risk_free_rate) / sigma_annual
            else:
                sharpe_ratio = np.nan
            
            # Information Ratio
            tracking_error = self.calculate_tracking_error(stock_returns, market_returns)
            if tracking_error != 0:
                excess_return = mean_annual - (market_returns.mean() * self.trading_days)
                information_ratio = excess_return / tracking_error
            else:
                information_ratio = np.nan
            
            # Sortino Ratio (SOR)
            downside_dev = self.calculate_downside_deviation(stock_returns, target_return=0)
            if downside_dev != 0:
                sortino_ratio = (mean_annual - risk_free_rate) / downside_dev
            else:
                sortino_ratio = np.nan
            
            # 获取股票名称
            stock_name = self.get_stock_name(stock)
            
            # 获取 CML 权重
            cml_weight = cml_weights.get(stock, np.nan) if stock in available_stocks else np.nan
            
            results.append({
                '股票号': stock,
                '股票名': stock_name,
                '均数 (%)': mean_annual * 100,
                'Skewness': skewness,
                'Kurtosis': kurtosis,
                'Sigma (%)': sigma_annual * 100,
                '与上证相关系数': correlation,
                'Beta': beta,
                'Treynor Ratio': treynor_ratio,
                'Sharpe Ratio': sharpe_ratio,
                'Information Ratio': information_ratio,
                'SOR Ratio': sortino_ratio,
                'CML Weight': cml_weight
            })
        
        return pd.DataFrame(results)

    def export_to_excel(self, results_df, stock_type="cn_blue_chip", case_name=None, output_dir="dataIntegrator/modelService/financialAnalysis/output/portfolio_metrics"):
        """
        导出结果到 Excel 文件（支持滚动回测结果 + 多个 Pivot 表格 + 折线图）
        完全参考 stock_metrics_analysis.py
        
        参数:
        - results_df: 分析结果 DataFrame
        - stock_type: 股票类型（用于文件名）
        - case_name: 案例名称（用于文件名，优先使用）
        - output_dir: 输出目录
        """
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"✅ 创建输出目录: {output_dir}")
        
        # 生成文件名（带资产组合名称和时间戳）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 确定文件名称前缀
        if case_name:
            name_prefix = case_name
        else:
            # 根据 stock_type 生成中文名称
            name_map = {
                'cn_blue_chip': '中国蓝筹股组合',
                'cn_tech': '中国科技股组合',
                'cn_consumer': '中国大消费组合',
                'cn_financial': '中国金融股组合',
                'cn_energy': '中国能源与制造业组合',
                'cn_custom': '中国自定义股票组合',
                'us_tech': '美国科技股',
                'us_finance': '美国金融股',
                'us_mixed': '美国混合股票',
                'us_custom': '美国自定义组合',
            }
            name_prefix = name_map.get(stock_type, stock_type)
        
        filename = f"股票指标滚动回测_{name_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 导出到 Excel
        logger.info(f"\n📄 导出到 Excel: {filepath}")
        
        # 定义需要生成 Pivot 表的指标列表
        pivot_metrics = [
            ('均数 (%)', '均数Pivot表'),
            ('Skewness', 'Skewness Pivot表'),
            ('Kurtosis', 'Kurtosis Pivot表'),
            ('Sigma (%)', 'Sigma Pivot表'),
            ('与上证相关系数', '相关系数Pivot表'),
            ('Beta', 'Beta Pivot表'),
            ('Treynor Ratio', 'Treynor Ratio Pivot表'),
            ('Sharpe Ratio', 'Sharpe Ratio Pivot表'),
            ('Information Ratio', 'Information Ratio Pivot表'),
            ('SOR Ratio', 'SOR Ratio Pivot表'),
            ('CML Weight', 'CML Weight Pivot表'),
        ]
        
        # 创建 Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: 主要数据表（完整指标）- 保留原始精度
            results_df.to_excel(writer, sheet_name='滚动回测结果', index=False)
            
            # 获取工作表并调整列宽
            worksheet = writer.sheets['滚动回测结果']
            
            # 设置列宽
            column_widths = {
                'A': 15,  # 回测日期
                'B': 15,  # 股票号
                'C': 15,  # 股票名
                'D': 12,  # 均数
                'E': 12,  # Skewness
                'F': 12,  # Kurtosis
                'G': 12,  # Sigma
                'H': 15,  # 相关系数
                'I': 10,  # Beta
                'J': 15,  # Treynor Ratio
                'K': 15,  # Sharpe Ratio
                'L': 18,  # Information Ratio
                'M': 12,  # SOR Ratio
                'N': 15,  # CML Weight
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Sheet 2-12: 各个指标的 Pivot 表格 + 折线图（4位小数）
            for idx, (metric_col, sheet_name) in enumerate(pivot_metrics, 2):
                logger.info(f"\n🔄 生成 {sheet_name}...")
                pivot_df = self._create_metric_pivot_table(results_df, metric_col)
                
                # 数值列保留4位小数
                numeric_cols = pivot_df.columns[pivot_df.columns != '回测日期']
                for col in numeric_cols:
                    pivot_df[col] = pivot_df[col].round(4)
                
                pivot_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 调整 Pivot 表格列宽
                pivot_worksheet = writer.sheets[sheet_name]
                pivot_worksheet.column_dimensions['A'].width = 15  # 回测日期列
                
                # 其他列（股票名）自动调整
                for col in pivot_worksheet.columns:
                    if col[0].column_letter != 'A':  # 跳过第一列
                        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                        pivot_worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 15)
                
                # 添加图表（CML Weight 使用堆积面积图，其他使用折线图）
                logger.info(f"   📊 正在添加图表...")
                if metric_col == 'CML Weight':
                    # CML Weight 使用堆积面积图
                    self._add_area_chart_to_sheet(pivot_worksheet, sheet_name, metric_col)
                else:
                    # 其他指标使用折线图
                    self._add_line_chart_to_sheet(pivot_worksheet, sheet_name, metric_col)
                
                logger.info(f"   ✅ {sheet_name} 创建完成")
        
        logger.info(f"\n✅ Excel 文件已保存: {filepath}")
        logger.info(f"   📊 Sheet 1: 滚动回测结果 (完整指标)")
        for idx, (_, sheet_name) in enumerate(pivot_metrics, 2):
            logger.info(f"   📊 Sheet {idx}: {sheet_name}")
        
        # 生成并保存 PDF
        pdf_filepath = self.export_pivot_charts_to_pdf(
            results_df, 
            pivot_metrics, 
            stock_type=stock_type,
            case_name=case_name,
            output_dir=output_dir,
            timestamp=timestamp
        )
        
        return filepath, pdf_filepath

    def _create_metric_pivot_table(self, results_df, metric_column):
        """
        创建指定指标的 Pivot 表格
        
        参数:
        - results_df: 原始结果 DataFrame
        - metric_column: 指标列名
        
        返回:
        - pivot_df: Pivot 后的 DataFrame，行=回测日期，列=股票名，值=指标值
        """
        # 选择需要的列：回测日期、股票名、指标值
        pivot_data = results_df[['回测日期', '股票名', metric_column]].copy()
        
        # 创建 Pivot 表格
        pivot_df = pivot_data.pivot_table(
            index='回测日期',
            columns='股票名',
            values=metric_column,
            aggfunc='mean'  # 如果有重复值，取平均
        )
        
        # 重置索引，让回测日期成为普通列
        pivot_df = pivot_df.reset_index()
        
        # 按日期排序
        pivot_df = pivot_df.sort_values('回测日期').reset_index(drop=True)
        
        return pivot_df

    def _add_line_chart_to_sheet(self, worksheet, sheet_name, metric_name):
        """
        在工作表中添加折线图（右侧图例显示股票名称）
        
        参数:
        - worksheet: openpyxl 工作表对象
        - sheet_name: 工作表名称
        - metric_name: 指标名称（用于图表标题）
        """
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import DateAxis, ChartLines
        from openpyxl.chart.legend import Legend
        
        # 获取数据范围
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row < 2 or max_col < 2:
            logger.warning(f"   ⚠️ {sheet_name} 数据不足，跳过图表生成")
            return
        
        # 创建折线图
        chart = LineChart()
        chart.title = f"{metric_name} 趋势图"
        chart.style = 10  # 使用样式10
        
        # 设置坐标轴标题
        chart.y_axis.title = metric_name
        chart.x_axis.title = "回测日期"
        
        # 设置X轴标签旋转（垂直显示，参考图中样式）
        chart.x_axis.tickLblPos = "low"  # 标签位置
        
        # 显示图例（右侧，显示股票名称）
        chart.legend = Legend()
        chart.legend.position = 'r'  # 'r' = right
        chart.legend.overlay = False  # 不覆盖图表区域
        
        # 设置图表大小（有图例需要留出空间）
        chart.width = 28  # 厘米
        chart.height = 15  # 厘米
        
        # X轴数据（日期列，从第2行开始，跳过标题）
        dates = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
        
        # 为每一只股票添加一条线（从第2列开始，跳过日期列）
        for col_idx in range(2, max_col + 1):
            # 获取该列的数据（包括标题行用于图例）
            data = Reference(worksheet, min_col=col_idx, min_row=1, max_row=max_row)
            
            # 添加到图表
            chart.add_data(data, titles_from_data=True)
        
        # 设置X轴标签
        chart.set_categories(dates)
        
        # 添加网格线（参考图中有网格线）
        chart.y_axis.majorGridlines = ChartLines()  # Y轴主网格线
        chart.x_axis.majorGridlines = ChartLines()  # X轴主网格线
        
        # 将图表添加到工作表（从A1开始，覆盖数据区域）
        anchor_cell = worksheet.cell(row=1, column=1)
        worksheet.add_chart(chart, anchor_cell.coordinate)
        
        logger.info(f"   ✅ 折线图已添加到 {sheet_name}（右侧图例）")
    
    def _add_area_chart_to_sheet(self, worksheet, sheet_name, metric_name):
        """
        在工作表中添加堆积面积图（用于 CML Weight）
        
        参数:
        - worksheet: openpyxl 工作表对象
        - sheet_name: 工作表名称
        - metric_name: 指标名称（用于图表标题）
        """
        from openpyxl.chart import AreaChart, Reference
        from openpyxl.chart.axis import ChartLines
        from openpyxl.chart.legend import Legend
        
        # 获取数据范围
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row < 2 or max_col < 2:
            logger.warning(f"   ⚠️ {sheet_name} 数据不足，跳过图表生成")
            return
        
        # 创建堆积面积图
        chart = AreaChart()
        chart.title = f"{metric_name} 趋势图"
        chart.style = 10  # 使用样式10
        chart.grouping = "stacked"  # 堆积模式
        
        # 设置坐标轴标题
        chart.y_axis.title = metric_name
        chart.x_axis.title = "回测日期"
        
        # 设置Y轴为百分比格式
        chart.y_axis.number_format = '0%'
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1  # 100%
        
        # 设置X轴标签位置
        chart.x_axis.tickLblPos = "low"
        
        # 显示图例（右侧）
        chart.legend = Legend()
        chart.legend.position = 'r'
        chart.legend.overlay = False
        
        # 设置图表大小
        chart.width = 28  # 厘米
        chart.height = 15  # 厘米
        
        # X轴数据（日期列，从第2行开始）
        dates = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
        
        # 为每一只股票添加一个面积层（从第2列开始）
        for col_idx in range(2, max_col + 1):
            data = Reference(worksheet, min_col=col_idx, min_row=1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
        
        # 设置X轴标签
        chart.set_categories(dates)
        
        # 添加网格线
        chart.y_axis.majorGridlines = ChartLines()
        chart.x_axis.majorGridlines = ChartLines()
        
        # 将图表添加到工作表
        anchor_cell = worksheet.cell(row=1, column=1)
        worksheet.add_chart(chart, anchor_cell.coordinate)
        
        logger.info(f"   ✅ 堆积面积图已添加到 {sheet_name}")
    
    def export_pivot_charts_to_pdf(self, results_df, pivot_metrics, stock_type="cn_blue_chip", 
                                    case_name=None, output_dir="dataIntegrator/modelService/financialAnalysis/output/portfolio_metrics",
                                    timestamp=None):
        """
        将 Pivot 图表导出为 PDF（参考 stock_metrics_analysis.py 的逻辑）
        PDF 里图片在前，居中，带标题，窄margin，然后就是下一张图
        
        参数:
        - results_df: 原始结果 DataFrame
        - pivot_metrics: Pivot 指标列表 [(metric_col, sheet_name), ...]
        - stock_type: 股票类型（用于文件名）
        - case_name: 案例名称（用于文件名，优先使用）
        - output_dir: PDF 输出目录
        - timestamp: 时间戳（用于文件名）
        
        返回:
        - pdf_filepath: PDF 文件路径
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import matplotlib.pyplot as plt
            from io import BytesIO
            
            # 确保输出目录存在
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 生成 PDF 文件名（与 Excel 同名）
            if timestamp is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 确定文件名称前缀
            if case_name:
                name_prefix = case_name
            else:
                name_map = {
                    'cn_blue_chip': '中国蓝筹股组合',
                    'cn_tech': '中国科技股组合',
                    'cn_consumer': '中国大消费组合',
                    'cn_financial': '中国金融股组合',
                    'cn_energy': '中国能源与制造业组合',
                    'cn_custom': '中国自定义股票组合',
                    'us_tech': '美国科技股',
                    'us_finance': '美国金融股',
                    'us_mixed': '美国混合股票',
                    'us_custom': '美国自定义组合',
                }
                name_prefix = name_map.get(stock_type, stock_type)
            
            pdf_filename = f"股票指标Pivot分析_{name_prefix}_{timestamp}.pdf"
            pdf_filepath = os.path.join(output_dir, pdf_filename)
            
            logger.info(f"\n📄 开始生成 PDF: {pdf_filepath}")
            
            # 注册中文字体（使用系统字体）
            try:
                # Windows 系统字体
                font_path = r"C:\Windows\Fonts\simhei.ttf"  # 黑体
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('SimHei', font_path))
                    chinese_font = 'SimHei'
                else:
                    chinese_font = 'Helvetica'  # fallback
            except:
                chinese_font = 'Helvetica'
            
            # 创建 PDF 文档（横向 A4，窄边距）
            doc = SimpleDocTemplate(
                pdf_filepath,
                pagesize=landscape(A4),
                rightMargin=0.5*cm,
                leftMargin=0.5*cm,
                topMargin=1*cm,
                bottomMargin=0.5*cm
            )
            
            # 定义样式
            styles = getSampleStyleSheet()
            
            # 标题样式
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                fontName=chinese_font,
                spaceAfter=8,
                alignment=1  # 居中
            )
            
            # 注释样式（专业分析文字）
            comment_style = ParagraphStyle(
                'CommentStyle',
                parent=styles['Normal'],
                fontSize=9,
                fontName=chinese_font,
                leading=12,
                spaceAfter=8,
                leftIndent=10,
                rightIndent=10,
                backColor=colors.lightyellow
            )
            
            # 构建 PDF 内容
            story = []
            
            # 遍历每个 Pivot 指标
            for idx, (metric_col, sheet_name) in enumerate(pivot_metrics, 1):
                logger.info(f"   📊 处理 {idx}/{len(pivot_metrics)}: {sheet_name}")
                
                # 创建 Pivot 数据
                pivot_df = self._create_metric_pivot_table(results_df, metric_col)
                
                # 数值列保疙4位小数
                numeric_cols = pivot_df.columns[pivot_df.columns != '回测日期']
                for col in numeric_cols:
                    pivot_df[col] = pivot_df[col].round(4)
                
                # 生成图表图片
                try:
                    chart_image = self._generate_chart_image(pivot_df, metric_col, sheet_name)
                    if chart_image:
                        # 添加标题
                        story.append(Paragraph(f"{sheet_name}", title_style))
                        story.append(Spacer(1, 0.3*cm))
                        
                        # 添加图片（居中）
                        img = Image(chart_image)
                        # 调整图片大小以适应横向页面
                        img.drawWidth = 27*cm  # 横向A4宽度约29.7cm，减去边距
                        img.drawHeight = 16*cm
                        story.append(img)
                        
                        # 添加专业分析评论
                        comment = self.generate_professional_comment(metric_col, pivot_df)
                        # 将换行符转换为 HTML 格式
                        comment_html = comment.replace('\n', '<br/>')
                        story.append(Paragraph(comment_html, comment_style))
                        story.append(Spacer(1, 0.5*cm))
                        
                        # 分页符（除了最后一个）
                        if idx < len(pivot_metrics):
                            story.append(PageBreak())
                except Exception as e:
                    logger.warning(f"   ⚠️ 图表生成失败: {e}")
                    continue
            
            # 生成 PDF
            doc.build(story)
            
            logger.info(f"✅ PDF 文件已保存: {pdf_filepath}")
            return pdf_filepath
        
        except ImportError as e:
            logger.error(f"❌ 缺少必要的库: {e}")
            logger.info("💡 请安装: pip install reportlab matplotlib")
            return None
        except Exception as e:
            logger.error(f"❌ PDF 生成失败: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def _generate_chart_image(self, pivot_df, metric_col, sheet_name):
        """
        生成图表图片（内存中的 PNG）
        
        参数:
        - pivot_df: Pivot 数据 DataFrame
        - metric_col: 指标列名
        - sheet_name: 工作表名称
        
        返回:
        - image_bytes: 图片字节流
        """
        import matplotlib.pyplot as plt
        
        try:
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
            plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
            
            # 准备数据
            dates = pivot_df['回测日期'].values
            numeric_cols = pivot_df.columns[pivot_df.columns != '回测日期']
            
            if len(numeric_cols) == 0 or len(dates) == 0:
                return None
            
            # 创建图表
            fig, ax = plt.subplots(figsize=(14, 7))
            
            # 根据指标类型选择图表类型
            if metric_col == 'CML Weight':
                # CML Weight 使用堆积面积图
                ax.stackplot(range(len(dates)), 
                           [pivot_df[col].fillna(0).values for col in numeric_cols],
                           labels=numeric_cols,
                           alpha=0.7)
            else:
                # 其他指标使用折线图
                for col in numeric_cols:
                    values = pivot_df[col].values
                    ax.plot(range(len(dates)), values, label=col, linewidth=1.5)
            
            # 设置标题和标签
            ax.set_title(f"{metric_col} 趋势图", fontsize=14, fontweight='bold')
            ax.set_xlabel('回测日期', fontsize=11)
            ax.set_ylabel(metric_col, fontsize=11)
            
            # 设置X轴刻度（只显示部分日期标签）
            step = max(1, len(dates) // 10)
            tick_positions = range(0, len(dates), step)
            tick_labels = [str(d)[4:] if isinstance(d, str) else str(d)[-6:] for d in dates]
            ax.set_xticks(tick_positions)
            ax.set_xticklabels([tick_labels[i] if i < len(tick_labels) else '' for i in tick_positions], 
                             rotation=45, ha='right', fontsize=8)
            
            # 添加图例（右侧）
            ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=8, framealpha=0.9)
            
            # 添加网格线
            ax.grid(True, linestyle='--', alpha=0.5)
            
            # 调整布局
            plt.tight_layout()
            
            # 保存到内存缓冲区
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig)
            
            return buf
        
        except Exception as e:
            logger.warning(f"图表生成异常: {e}")
            return None
    
    def generate_professional_comment(self, metric_col, pivot_df):
        """
        为指定指标生成专业的分析评论（委托给 PortfolioMetricsAnalysisReport）
        
        参数:
        - metric_col: 指标列名
        - pivot_df: Pivot 数据 DataFrame
        
        返回:
        - comment: 专业分析评论文本
        """
        from dataIntegrator.modelService.financialAnalysis.PortfolioMetricsAnalysisReport import PortfolioMetricsAnalysisReport
        
        report_generator = PortfolioMetricsAnalysisReport()
        return report_generator.generate_professional_comment(metric_col, pivot_df)


if __name__ == "__main__":
    from dataIntegrator.modelService.financialAnalysis.PortfolioMetricsAnalysisTest import PortfolioMetricsAnalysisTest

    # 创建分析器
    analyzer = PortfolioMetricsAnalysis()
    
    # 执行滚动回测分析（完全参考 stock_metrics_analysis.py 的 main 函数）
    results_df = analyzer.analyze_stocks_rolling(
        stock_type="cn_blue_chip",  # 股票类型
        start_date_fixed=None,  # 自动计算
        end_date_start='20251201',  # 滚动起始日期
        end_date_end=CommonParameters.today,  # 滚动结束日期（今天）
        window_days=360,  # 回看窗口360天
        risk_free_rate=None  # 自动获取
    )
    
    # 打印前几行结果
    logger.info("\n📊 分析结果预览（前10条）:")
    logger.info("-" * 80)
    print(results_df.head(10).to_string(index=False))
    logger.info("-" * 80)
    
    # 导出到 Excel
    excel_path = analyzer.export_to_excel(results_df)
    
    logger.info("\n" + "=" * 80)
    logger.info("🎉 所有任务完成！")
    logger.info(f"📁 Excel 文件: {excel_path}")
    logger.info(f"📊 总记录数: {len(results_df)}")
    logger.info("=" * 80)
